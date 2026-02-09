"""YJK柱脚内力处理工具业务逻辑"""

import os
import pandas as pd
from PySide6.QtWidgets import QFileDialog


class YJKColumnForceLogic:
    """YJK柱脚内力处理工具业务逻辑"""
    
    def __init__(self):
        """初始化业务逻辑"""
        pass
    
    def process_pressure(self, file_path: str):
        """处理压力数据
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            dict: 处理结果，包含success、original_rows、removed_rows、final_rows、save_path、format、error等字段
        """
        try:
            # 尝试读取工作表
            df = self._read_excel_file(file_path)
            
            # 处理数据
            processed_data = self._process_data(df, mode="pressure", file_path=file_path)
            
            # 导出数据
            result = self._export_data(
                processed_data["df"], 
                file_path, 
                "压力",
                processed_data["original_rows"],
                processed_data["removed_rows"],
                processed_data["final_rows"],
                processed_data.get("vlookup_status", "未执行")
            )
            
            return result
        
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def process_tension(self, file_path: str):
        """处理拉力数据
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            dict: 处理结果，包含success、original_rows、removed_rows、final_rows、save_path、format、error等字段
        """
        try:
            # 尝试读取工作表
            df = self._read_excel_file(file_path)
            
            # 处理数据
            processed_data = self._process_data(df, mode="tension", file_path=file_path)
            
            # 导出数据
            result = self._export_data(
                processed_data["df"], 
                file_path, 
                "拉力",
                processed_data["original_rows"],
                processed_data["removed_rows"],
                processed_data["final_rows"],
                processed_data.get("vlookup_status", "未执行")
            )
            
            return result
        
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def process_all(self, file_path: str):
        """处理全部柱底内力
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            dict: 处理结果，包含success、original_rows、removed_rows、final_rows、save_path、format、error等字段
        """
        try:
            # 尝试读取工作表
            df = self._read_excel_file(file_path)
            
            # 处理数据
            processed_data = self._process_data(df, mode="all", file_path=file_path)
            
            # 导出数据
            result = self._export_data(
                processed_data["df"], 
                file_path, 
                "全部柱底内力",
                processed_data["original_rows"],
                processed_data["removed_rows"],
                processed_data["final_rows"],
                processed_data.get("vlookup_status", "未执行")
            )
            
            return result
        
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def _read_excel_file(self, file_path: str):
        """读取Excel文件
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            pd.DataFrame: 读取的数据
        """
        df = None
        
        try:
            # 首先尝试直接使用openpyxl读取，绕过pandas的限制
            from openpyxl import load_workbook
            
            # 尝试使用不同的openpyxl参数组合
            openpyxl_params = [
                # 基本只读模式
                {"read_only": True, "data_only": True},
                # 常规模式
                {"read_only": False, "data_only": True},
                # 只读但不读取数据值
                {"read_only": True, "data_only": False},
                # 不使用只读模式
                {"read_only": False, "data_only": False}
            ]
            
            wb = None
            for params in openpyxl_params:
                try:
                    wb = load_workbook(file_path, **params)
                    break
                except Exception:
                    continue
            
            if wb:
                # 读取工作表
                if "基本组合内力" in wb.sheetnames:
                    ws = wb["基本组合内力"]
                    
                    # 读取所有数据
                    data = []
                    for row in ws.iter_rows(values_only=True):
                        data.append(list(row))
                    
                    # 转换为DataFrame，不设置字母列名
                    df = pd.DataFrame(data)
                    wb.close()
                else:
                    raise ValueError("工作表'基本组合内力'不存在")
        
        except Exception as openpyxl_e:
            # 如果openpyxl读取失败，尝试使用pandas的各种选项
            read_options = [
                # 常规读取
                {"header": 0},
                # 常规读取，无表头
                {"header": None},
                # 使用openpyxl引擎
                {"header": 0, "engine": "openpyxl"},
                {"header": None, "engine": "openpyxl"},
                # 使用xlrd引擎（如果可用）
                {"header": 0, "engine": "xlrd"},
                {"header": None, "engine": "xlrd"},
                # 使用openpyxl引擎，数据类型选项
                {"header": 0, "engine": "openpyxl", "dtype": str},
                {"header": None, "engine": "openpyxl", "dtype": str}
            ]
            
            for options in read_options:
                try:
                    df = pd.read_excel(file_path,
                                    sheet_name="基本组合内力",
                                    **options)
                    break
                except Exception:
                    continue
        
        if df is None:
            # 尝试使用最基本的方式读取
            df = pd.read_excel(file_path,
                            sheet_name="基本组合内力",
                            header=None,
                            engine="openpyxl")
            if df.shape[1] >= 12:
                pass
            else:
                raise ValueError(f"工作表列数不足（需要至少12列），当前列数: {df.shape[1]}")
        
        return df
    
    def _process_data(self, df: pd.DataFrame, mode: str = "all", file_path: str = None):
        """处理数据
        
        Args:
            df: 原始数据
            mode: 处理模式，可选值：pressure（压力）、tension（拉力）、all（全部）
            file_path: Excel文件路径，用于读取"构件基本信息"工作表
            
        Returns:
            dict: 处理后的数据和统计信息
        """
        # 检查数据格式，确保保留原有列数和第一行
        if df.shape[0] < 1:
            raise ValueError("文件中没有数据")
        
        # 检查列数是否足够（至少12列，对应A-L）
        if df.shape[1] < 12:
            raise ValueError(f"工作表列数不足（需要至少12列），当前列数: {df.shape[1]}")
        
        # 分离第一行（表头）和数据部分
        header_row = df.iloc[0:1].copy()  # 保留第一行
        data_rows = df.iloc[1:].copy()  # 数据部分从第二行开始
        
        # 记录原始数据行数
        original_data_rows = len(data_rows)
        
        # ========== VLOOKUP处理（移到数据处理最开始） ==========
        print(f"开始VLOOKUP处理，文件路径: {file_path}")
        if file_path:
            try:
                print(f"尝试读取Excel文件: {file_path}")
                # 读取"构件基本信息"工作表
                from openpyxl import load_workbook
                wb = load_workbook(file_path, read_only=True, data_only=True)
                print(f"成功打开Excel文件")
                
                if "构件基本信息" in wb.sheetnames:
                    ws = wb["构件基本信息"]
                    
                    # 读取构件基本信息数据
                    component_data = []
                    for row in ws.iter_rows(values_only=True):
                        component_data.append(list(row))
                    
                    # 转换为DataFrame
                    component_df = pd.DataFrame(component_data)
                    
                    # 打印构件基本信息表的结构
                    print(f"构件基本信息表列数: {component_df.shape[1]}")
                    print(f"构件基本信息表前5行:")
                    print(component_df.head())
                    
                    # 确保构件基本信息有足够的列
                    if component_df.shape[1] >= 7:  # 至少有G列（索引6）
                        # 打印D列的前10个值
                        print(f"D列（索引3）前10个值: {list(component_df.iloc[:, 3].head(10))}")
                        # 打印G列的前10个值
                        print(f"G列（索引6）前10个值: {list(component_df.iloc[:, 6].head(10))}")
                        
                        # 提取D列（索引3）和G列（索引6）
                        component_lookup = component_df.iloc[:, [3, 6]]  # D列（索引3）和G列（索引6）
                        
                        # 移除空行
                        component_lookup = component_lookup.dropna(subset=[3])
                        
                        # 创建查找字典：键是构件基本信息表的D列，值是构件基本信息表的G列
                        lookup_dict = dict(zip(component_lookup.iloc[:, 0], component_lookup.iloc[:, 1]))
                        
                        # 打印查找字典信息
                        print(f"查找字典大小: {len(lookup_dict)}")
                        print(f"前5个查找项: {list(lookup_dict.items())[:5]}")
                        
                        # 添加O列（第15列），使用VLOOKUP逻辑
                        # 对应Excel函数：VLOOKUP(D当前行, 构件基本信息!D:G, 4, 0)
                        if data_rows.shape[1] >= 4:  # 确保有D列（索引3）
                            # 打印D列的前5个值
                            print(f"基本组合内力表D列前5个值: {list(data_rows.iloc[:, 3].head())}")
                            
                            # 应用VLOOKUP逻辑
                            def vlookup_value(d_value):
                                result = lookup_dict.get(d_value, "")
                                print(f"查找值: {d_value}, 结果: {result}")
                                return result
                            
                            # 添加O列数据
                            o_column_data = data_rows.iloc[:, 3].apply(vlookup_value)
                            
                            # 打印O列的前5个值
                            print(f"O列前5个值: {list(o_column_data.head())}")
                            
                            # 将O列数据添加到data_rows
                            data_rows[14] = o_column_data  # 14是O列的索引（从0开始）
                            
                            # 更新表头
                            if header_row.shape[1] <= 14:
                                header_row[14] = "构件型号"  # 添加O列表头
                            
                            # 打印最终数据行数
                            print(f"处理后的数据行数: {len(data_rows)}")
                else:
                    print("工作表'构件基本信息'不存在")
                
                wb.close()
            except Exception as e:
                # 如果读取构件基本信息失败，不影响主流程
                print(f"VLOOKUP处理异常: {str(e)}")
                pass
        # ========== VLOOKUP处理结束 ==========
        
        # 转换B~L列为数值（对应数字索引1~11，即第2列到第12列）
        data_rows.iloc[:, 1:12] = data_rows.iloc[:, 1:12].apply(pd.to_numeric, errors='coerce')
        
        # 删除F列（索引5）值为1的行
        data_rows = data_rows[data_rows.iloc[:, 5] != 1]
        removed_f1_rows = original_data_rows - len(data_rows)
        
        # 按K列（索引10）倒序排序
        data_rows = data_rows.sort_values(by=data_rows.columns[10], ascending=False)
        
        # 根据模式过滤数据
        if mode == "pressure":
            # 仅保留K列<0的行（压力）
            data_rows = data_rows[data_rows.iloc[:, 10] < 0]
        elif mode == "tension":
            # 仅保留K列>0的行（拉力）
            data_rows = data_rows[data_rows.iloc[:, 10] > 0]
        # 否则保留所有数据（mode="all"）
        
        final_rows = len(data_rows)
        
        if final_rows == 0:
            if mode == "pressure":
                raise ValueError("没有找到K列<0的数据")
            elif mode == "tension":
                raise ValueError("没有找到K列>0的数据")
            else:
                raise ValueError("没有找到符合条件的数据")
        
        # 打印数据处理后的列数和O列状态
        print(f"处理后的数据列数: {data_rows.shape[1]}")
        print(f"O列是否存在: {14 in data_rows.columns}")
        if 14 in data_rows.columns:
            print(f"O列非空值数量: {data_rows[14].notna().sum()}")
            print(f"O列前5个值: {list(data_rows[14].head())}")
        
        # 重新组合表头和处理后的数据
        processed_df = pd.concat([header_row, data_rows], ignore_index=True)
        
        # 打印最终DataFrame的列数
        print(f"最终DataFrame列数: {processed_df.shape[1]}")
        if processed_df.shape[1] > 14:
            print(f"最终O列前5个值: {list(processed_df.iloc[:, 14].head())}")
        
        # 检查O列是否存在且有数据
        vlookup_status = "未执行"
        if processed_df.shape[1] > 14:
            o_column_values = processed_df.iloc[1:, 14]  # 排除表头行
            non_empty_count = o_column_values.notna().sum()
            if non_empty_count > 0:
                vlookup_status = f"成功，找到 {non_empty_count} 个匹配项"
            else:
                vlookup_status = "执行了但未找到匹配项"
        
        return {
            "df": processed_df,
            "original_rows": original_data_rows,
            "removed_rows": removed_f1_rows,
            "final_rows": final_rows,
            "vlookup_status": vlookup_status
        }
    
    def _export_data(self, df: pd.DataFrame, original_file_path: str, data_type: str, 
                    original_rows: int, removed_rows: int, final_rows: int, vlookup_status: str = "未执行"):
        """导出数据
        
        Args:
            df: 处理后的数据
            original_file_path: 原始文件路径
            data_type: 数据类型（压力、拉力、全部柱底内力）
            original_rows: 原始数据行数
            removed_rows: 删除的行数
            final_rows: 最终数据行数
            vlookup_status: VLOOKUP处理状态
            
        Returns:
            dict: 导出结果
        """
        # 弹出格式选择对话框
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QLabel, QRadioButton, QButtonGroup, QPushButton, QFileDialog, QMessageBox
        from PySide6.QtCore import Qt
        
        dialog = QDialog()
        dialog.setWindowTitle("选择导出格式")
        dialog.setFixedSize(300, 150)
        
        layout = QVBoxLayout(dialog)
        
        label = QLabel("请选择导出文件格式:")
        label.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)
        
        # 按钮组
        button_group = QButtonGroup(dialog)
        
        # Excel格式
        excel_radio = QRadioButton("Excel格式 (.xlsx)")
        excel_radio.setChecked(True)
        button_group.addButton(excel_radio)
        layout.addWidget(excel_radio)
        
        # TXT格式
        txt_radio = QRadioButton("文本格式 (.txt)")
        button_group.addButton(txt_radio)
        layout.addWidget(txt_radio)
        
        # 确定按钮
        ok_button = QPushButton("确定")
        ok_button.clicked.connect(dialog.accept)
        layout.addWidget(ok_button)
        
        # 居中显示
        dialog.exec_()
        
        # 获取选择的格式
        if excel_radio.isChecked():
            export_format = "xlsx"
        else:
            export_format = "txt"
        
        # 检查是否有O列数据（索引14）
        has_o_column = df.shape[1] > 14
        
        if has_o_column:
            # 获取O列数据（排除表头）
            o_column_data = df.iloc[1:, 14]
            # 获取唯一的O列值（排除空值）
            unique_o_values = o_column_data[o_column_data.notna() & (o_column_data != '')].unique()
            
            if len(unique_o_values) > 0:
                # 根据O列数据分组导出
                save_paths = []
                base_dir = os.path.dirname(original_file_path)
                base_name = os.path.splitext(os.path.basename(original_file_path))[0]
                
                for o_value in unique_o_values:
                    # 过滤出当前O列值的数据
                    # 保留表头
                    header = df.iloc[0:1].copy()
                    # 过滤数据行
                    filtered_data = df.iloc[1:].copy()
                    filtered_data = filtered_data[filtered_data.iloc[:, 14] == o_value]
                    # 重新组合表头和过滤后的数据
                    grouped_df = pd.concat([header, filtered_data], ignore_index=True)
                    
                    # 计算当前分组的行数
                    grouped_rows = len(filtered_data)
                    
                    if grouped_rows > 0:
                        # 生成文件名
                        safe_o_value = str(o_value).replace('×', 'x').replace('/', '').replace('\\', '')
                        if export_format == "xlsx":
                            file_name = f"{base_name}{data_type}_{safe_o_value}.xlsx"
                        else:
                            file_name = f"{base_name}{data_type}_{safe_o_value}.txt"
                        
                        save_path = os.path.join(base_dir, file_name)
                        save_paths.append(save_path)
                        
                        if export_format == "xlsx":
                            # Excel格式导出
                            # 添加统计信息
                            summary = pd.DataFrame({
                                '统计项': ['原始数据行数', '删除F=1行数', f'最终{data_type}行数', 'VLOOKUP处理状态', '构件型号'],
                                '数值': [original_rows, removed_rows, grouped_rows, vlookup_status, o_value]
                            })
                            
                            # 使用ExcelWriter保存多个工作表，不写入列名，保留原始数据格式
                            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                                # 写入数据
                                grouped_df.to_excel(writer, index=False, header=False, sheet_name=f"{data_type}数据")
                                summary.to_excel(writer, index=False, sheet_name="处理统计")
                                
                                # 设置单元格样式：水平居中、垂直居中
                                from openpyxl.styles import Alignment
                                
                                # 处理数据工作表
                                ws = writer.sheets[f"{data_type}数据"]
                                alignment = Alignment(horizontal="center", vertical="center")
                                
                                # 遍历所有单元格并设置对齐方式
                                for row in ws.iter_rows():
                                    for cell in row:
                                        cell.alignment = alignment
                                
                                # 处理处理统计工作表
                                ws_summary = writer.sheets["处理统计"]
                                for row in ws_summary.iter_rows():
                                    for cell in row:
                                        cell.alignment = alignment
                        else:
                            # TXT格式导出
                            # 设置固定宽度，实现水平居中效果
                            column_widths = [15] * grouped_df.shape[1]  # 假设每列宽度为15个字符
                            
                            # 打开文件并写入数据
                            with open(save_path, 'w', encoding='utf-8') as f:
                                # 写入数据，使用固定宽度格式
                                for _, row in grouped_df.iterrows():
                                    # 格式化每行数据，使用固定宽度，居中对齐
                                    formatted_row = "".join([f"{str(cell):^{width}}" for cell, width in zip(row, column_widths)])
                                    f.write(formatted_row + '\n')
                
                if save_paths:
                    QMessageBox.information(None, "导出成功", f"已成功导出 {len(save_paths)} 个文件")
                    return {
                        "success": True,
                        "original_rows": original_rows,
                        "removed_rows": removed_rows,
                        "final_rows": final_rows,
                        "save_path": save_paths[0],
                        "save_paths": save_paths,
                        "format": export_format
                    }
        
        # 如果没有O列数据或分组失败，使用原有的导出逻辑
        # 根据选择的格式导出文件
        if export_format == "xlsx":
            # Excel格式
            default_name = os.path.splitext(os.path.basename(original_file_path))[0] + f"{data_type}.xlsx"
            file_types = "Excel文件 (*.xlsx);;所有文件 (*.*)"
        else:
            # TXT格式
            default_name = os.path.splitext(os.path.basename(original_file_path))[0] + f"{data_type}.txt"
            file_types = "文本文件 (*.txt);;所有文件 (*.*)"
        
        save_path, _ = QFileDialog.getSaveFileName(
            None, f"保存{data_type}数据文件",
            os.path.join(os.path.dirname(original_file_path), default_name),
            file_types)
        
        if not save_path:
            raise ValueError("用户取消了保存操作")
        
        if export_format == "xlsx":
            # Excel格式导出
            # 添加统计信息
            summary = pd.DataFrame({
                '统计项': ['原始数据行数', '删除F=1行数', f'最终{data_type}行数', 'VLOOKUP处理状态'],
                '数值': [original_rows, removed_rows, final_rows, vlookup_status]
            })
            
            # 使用ExcelWriter保存多个工作表，不写入列名，保留原始数据格式
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                # 写入数据
                df.to_excel(writer, index=False, header=False, sheet_name=f"{data_type}数据")
                summary.to_excel(writer, index=False, sheet_name="处理统计")
                
                # 设置单元格样式：水平居中、垂直居中
                from openpyxl.styles import Alignment
                
                # 处理数据工作表
                ws = writer.sheets[f"{data_type}数据"]
                alignment = Alignment(horizontal="center", vertical="center")
                
                # 遍历所有单元格并设置对齐方式
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = alignment
                
                # 处理处理统计工作表
                ws_summary = writer.sheets["处理统计"]
                for row in ws_summary.iter_rows():
                    for cell in row:
                        cell.alignment = alignment
        else:
            # TXT格式导出
            # 设置固定宽度，实现水平居中效果
            column_widths = [15] * df.shape[1]  # 假设每列宽度为15个字符
            
            # 打开文件并写入数据
            with open(save_path, 'w', encoding='utf-8') as f:
                # 写入数据，使用固定宽度格式
                for _, row in df.iterrows():
                    # 格式化每行数据，使用固定宽度，居中对齐
                    formatted_row = "".join([f"{str(cell):^{width}}" for cell, width in zip(row, column_widths)])
                    f.write(formatted_row + '\n')
        
        return {
            "success": True,
            "original_rows": original_rows,
            "removed_rows": removed_rows,
            "final_rows": final_rows,
            "save_path": save_path,
            "format": export_format
        }
    
    def export_explorer_data(self, file_path: str, export_type: str):
        """导出探索者数据
        
        Args:
            file_path: Excel文件路径
            export_type: 导出类型，可选值：pressure（压力）、tension（拉力）、all（全部）
            
        Returns:
            dict: 处理结果，包含success、original_rows、removed_rows、final_rows、save_path、format、error等字段
        """
        try:
            # 尝试读取工作表
            df = self._read_excel_file(file_path)
            
            # 检查数据格式，确保保留原有列数和第一行
            if df.shape[0] < 1:
                raise ValueError("文件中没有数据")
            
            # 检查列数是否足够（至少14列，对应A-N）
            if df.shape[1] < 14:
                raise ValueError(f"工作表列数不足（需要至少14列），当前列数: {df.shape[1]}")
            
            # 分离第一行（表头）和数据部分
            header_row = df.iloc[0:1].copy()  # 保留第一行
            data_rows = df.iloc[1:].copy()  # 数据部分从第二行开始
            
            # 记录原始数据行数
            original_data_rows = len(data_rows)
            
            # ========== VLOOKUP处理（移到数据处理最开始） ==========
            print(f"开始VLOOKUP处理，文件路径: {file_path}")
            vlookup_status = "未执行"
            o_column_data = None
            
            try:
                print(f"尝试读取Excel文件: {file_path}")
                # 读取"构件基本信息"工作表
                from openpyxl import load_workbook
                wb = load_workbook(file_path, read_only=True, data_only=True)
                print(f"成功打开Excel文件")
                
                if "构件基本信息" in wb.sheetnames:
                    ws = wb["构件基本信息"]
                    
                    # 读取构件基本信息数据
                    component_data = []
                    for row in ws.iter_rows(values_only=True):
                        component_data.append(list(row))
                    
                    # 转换为DataFrame
                    component_df = pd.DataFrame(component_data)
                    
                    # 打印构件基本信息表的结构
                    print(f"构件基本信息表列数: {component_df.shape[1]}")
                    print(f"构件基本信息表前5行:")
                    print(component_df.head())
                    
                    # 确保构件基本信息有足够的列
                    if component_df.shape[1] >= 7:  # 至少有G列（索引6）
                        # 打印D列的前10个值
                        print(f"D列（索引3）前10个值: {list(component_df.iloc[:, 3].head(10))}")
                        # 打印G列的前10个值
                        print(f"G列（索引6）前10个值: {list(component_df.iloc[:, 6].head(10))}")
                        
                        # 提取D列（索引3）和G列（索引6）
                        component_lookup = component_df.iloc[:, [3, 6]]  # D列（索引3）和G列（索引6）
                        
                        # 移除空行
                        component_lookup = component_lookup.dropna(subset=[3])
                        
                        # 创建查找字典：键是构件基本信息表的D列，值是构件基本信息表的G列
                        lookup_dict = dict(zip(component_lookup.iloc[:, 0], component_lookup.iloc[:, 1]))
                        
                        # 打印查找字典信息
                        print(f"查找字典大小: {len(lookup_dict)}")
                        print(f"前5个查找项: {list(lookup_dict.items())[:5]}")
                        
                        # 添加O列（第15列），使用VLOOKUP逻辑
                        # 对应Excel函数：VLOOKUP(D当前行, 构件基本信息!D:G, 4, 0)
                        if data_rows.shape[1] >= 4:  # 确保有D列（索引3）
                            # 打印D列的前5个值
                            print(f"基本组合内力表D列前5个值: {list(data_rows.iloc[:, 3].head())}")
                            
                            # 应用VLOOKUP逻辑
                            def vlookup_value(d_value):
                                result = lookup_dict.get(d_value, "")
                                print(f"查找值: {d_value}, 结果: {result}")
                                return result
                            
                            # 添加O列数据
                            o_column_data = data_rows.iloc[:, 3].apply(vlookup_value)
                            
                            # 打印O列的前5个值
                            print(f"O列前5个值: {list(o_column_data.head())}")
                            
                            # 将O列数据添加到data_rows
                            data_rows[14] = o_column_data  # 14是O列的索引（从0开始）
                            
                            # 更新表头
                            if header_row.shape[1] <= 14:
                                header_row[14] = "构件型号"  # 添加O列表头
                            
                            # 打印最终数据行数
                            print(f"处理后的数据行数: {len(data_rows)}")
                            vlookup_status = "成功"
                else:
                    print("工作表'构件基本信息'不存在")
                
                wb.close()
            except Exception as e:
                # 如果读取构件基本信息失败，不影响主流程
                print(f"VLOOKUP处理异常: {str(e)}")
                pass
            # ========== VLOOKUP处理结束 ==========
            
            # 转换B~L列为数值（对应数字索引1~11，即第2列到第12列）
            data_rows.iloc[:, 1:12] = data_rows.iloc[:, 1:12].apply(pd.to_numeric, errors='coerce')
            
            # 删除F列（索引5）值为1的行
            data_rows = data_rows[data_rows.iloc[:, 5] != 1]
            removed_f1_rows = original_data_rows - len(data_rows)
            
            # 按K列（索引10）倒序排序
            data_rows = data_rows.sort_values(by=data_rows.columns[10], ascending=False)
            
            # 根据导出类型过滤数据
            if export_type == "pressure":
                # 仅保留K列<0的行（压力）
                data_rows = data_rows[data_rows.iloc[:, 10] < 0]
                filtered_rows = len(data_rows)
                export_suffix = "压力"
            elif export_type == "tension":
                # 仅保留K列>0的行（拉力）
                data_rows = data_rows[data_rows.iloc[:, 10] > 0]
                filtered_rows = len(data_rows)
                export_suffix = "拉力"
            else:
                # 全部内力
                filtered_rows = len(data_rows)
                export_suffix = "全部内力"
            
            if filtered_rows == 0:
                raise ValueError(f"没有找到符合条件的数据")
            
            # 弹出格式选择对话框
            from PySide6.QtWidgets import QDialog, QVBoxLayout, QLabel, QRadioButton, QButtonGroup, QPushButton, QFileDialog, QMessageBox
            from PySide6.QtCore import Qt
            
            dialog = QDialog()
            dialog.setWindowTitle("选择导出格式")
            dialog.setFixedSize(300, 150)
            
            layout = QVBoxLayout(dialog)
            
            label = QLabel("请选择导出文件格式:")
            label.setAlignment(Qt.AlignCenter)
            layout.addWidget(label)
            
            # 按钮组
            button_group = QButtonGroup(dialog)
            
            # Excel格式
            excel_radio = QRadioButton("Excel格式 (.xlsx)")
            excel_radio.setChecked(True)
            button_group.addButton(excel_radio)
            layout.addWidget(excel_radio)
            
            # TXT格式
            txt_radio = QRadioButton("文本格式 (.txt)")
            button_group.addButton(txt_radio)
            layout.addWidget(txt_radio)
            
            # 确定按钮
            ok_button = QPushButton("确定")
            ok_button.clicked.connect(dialog.accept)
            layout.addWidget(ok_button)
            
            # 居中显示
            dialog.exec_()
            
            # 获取选择的格式
            if excel_radio.isChecked():
                export_format = "xlsx"
            else:
                export_format = "txt"
            
            # 检查是否有O列数据（索引14）
            has_o_column = '14' in data_rows.columns or 14 in data_rows.columns
            
            if has_o_column:
                # 获取O列数据（排除表头）
                o_column_data = data_rows.iloc[:, 14]
                # 获取唯一的O列值（排除空值）
                unique_o_values = o_column_data[o_column_data.notna() & (o_column_data != '')].unique()
                
                if len(unique_o_values) > 0:
                    # 根据O列数据分组导出
                    save_paths = []
                    base_dir = os.path.dirname(file_path)
                    base_name = os.path.splitext(os.path.basename(file_path))[0]
                    
                    for o_value in unique_o_values:
                        # 过滤出当前O列值的数据
                        filtered_data = data_rows[data_rows.iloc[:, 14] == o_value]
                        current_filtered_rows = len(filtered_data)
                        
                        if current_filtered_rows > 0:
                            # 构建探索者格式数据
                            # 探索者表头
                            explorer_header = ["序号", "描述", "轴力Nz", "剪力Vx", "剪力Vy", "弯矩Mx", "弯矩My", "是否抗震"]
                            
                            # 构建数据行
                            explorer_data = [explorer_header]
                            for idx, (_, row) in enumerate(filtered_data.iterrows(), 1):
                                explorer_row = [
                                    idx,  # 序号
                                    f"组合工况{idx}",  # 描述
                                    row.iloc[10],  # 轴力Nz (原K列)
                                    row.iloc[9],   # 剪力Vx (原J列)
                                    row.iloc[8],   # 剪力Vy (原I列)
                                    row.iloc[7],   # 弯矩Mx (原H列)
                                    row.iloc[6],   # 弯矩My (原G列)
                                    "否"           # 是否抗震
                                ]
                                explorer_data.append(explorer_row)
                            
                            # 转换为DataFrame
                            explorer_df = pd.DataFrame(explorer_data)
                            
                            # 生成文件名
                            safe_o_value = str(o_value).replace('×', 'x').replace('/', '').replace('\\', '')
                            file_name = f"{base_name}探索者{export_suffix}_{safe_o_value}.{export_format}"
                            save_path = os.path.join(base_dir, file_name)
                            save_paths.append(save_path)
                            
                            if export_format == "xlsx":
                                # Excel格式导出
                                # 使用ExcelWriter保存数据，写入列名
                                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                                    # 写入数据
                                    explorer_df.to_excel(writer, index=False, header=False, sheet_name=f"探索者{export_suffix}数据")
                                    
                                    # 设置单元格样式：水平居中、垂直居中
                                    from openpyxl.styles import Alignment
                                    
                                    # 处理数据工作表
                                    ws = writer.sheets[f"探索者{export_suffix}数据"]
                                    alignment = Alignment(horizontal="center", vertical="center")
                                    
                                    # 遍历所有单元格并设置对齐方式
                                    for row in ws.iter_rows():
                                        for cell in row:
                                            cell.alignment = alignment
                            else:
                                # TXT格式导出
                                # 设置固定宽度
                                column_widths = [8, 20, 15, 15, 15, 15, 15, 10]  # 各列宽度
                                
                                # 打开文件并写入数据
                                with open(save_path, 'w', encoding='utf-8') as f:
                                    for _, row in explorer_df.iterrows():
                                        formatted_row = "".join([f"{str(cell):^{width}}" for cell, width in zip(row, column_widths)])
                                        f.write(formatted_row + '\n')
                    
                    if save_paths:
                        QMessageBox.information(None, "导出成功", f"已成功导出 {len(save_paths)} 个文件")
                        return {
                            "success": True,
                            "original_rows": original_data_rows,
                            "removed_rows": removed_f1_rows,
                            "final_rows": filtered_rows,
                            "save_path": save_paths[0],
                            "save_paths": save_paths,
                            "format": export_format
                        }
            
            # 如果没有O列数据或分组失败，使用原有的导出逻辑
            # 构建探索者格式数据
            # 探索者表头
            explorer_header = ["序号", "描述", "轴力Nz", "剪力Vx", "剪力Vy", "弯矩Mx", "弯矩My", "是否抗震"]
            
            # 构建数据行
            explorer_data = [explorer_header]
            for idx, (_, row) in enumerate(data_rows.iterrows(), 1):
                explorer_row = [
                    idx,  # 序号
                    f"组合工况{idx}",  # 描述
                    row.iloc[10],  # 轴力Nz (原K列)
                    row.iloc[9],   # 剪力Vx (原J列)
                    row.iloc[8],   # 剪力Vy (原I列)
                    row.iloc[7],   # 弯矩Mx (原H列)
                    row.iloc[6],   # 弯矩My (原G列)
                    "否"           # 是否抗震
                ]
                explorer_data.append(explorer_row)
            
            # 转换为DataFrame
            explorer_df = pd.DataFrame(explorer_data)
            
            # 根据选择的格式导出文件
            default_name = os.path.splitext(os.path.basename(file_path))[0] + f"探索者{export_suffix}.{export_format}"
            file_types = f"{'Excel' if export_format == 'xlsx' else '文本'}文件 (*.{export_format});;所有文件 (*.*)"
            
            save_path, _ = QFileDialog.getSaveFileName(
                None, f"保存探索者{export_suffix}数据文件",
                os.path.join(os.path.dirname(file_path), default_name),
                file_types)
            
            if not save_path:
                raise ValueError("用户取消了保存操作")
            
            if export_format == "xlsx":
                # Excel格式导出
                # 使用ExcelWriter保存数据，写入列名
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    # 写入数据
                    explorer_df.to_excel(writer, index=False, header=False, sheet_name=f"探索者{export_suffix}数据")
                    
                    # 设置单元格样式：水平居中、垂直居中
                    from openpyxl.styles import Alignment
                    
                    # 处理数据工作表
                    ws = writer.sheets[f"探索者{export_suffix}数据"]
                    alignment = Alignment(horizontal="center", vertical="center")
                    
                    # 遍历所有单元格并设置对齐方式
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.alignment = alignment
            else:
                # TXT格式导出
                # 设置固定宽度
                column_widths = [8, 20, 15, 15, 15, 15, 15, 10]  # 各列宽度
                
                # 打开文件并写入数据
                with open(save_path, 'w', encoding='utf-8') as f:
                    for _, row in explorer_df.iterrows():
                        formatted_row = "".join([f"{str(cell):^{width}}" for cell, width in zip(row, column_widths)])
                        f.write(formatted_row + '\n')
            
            return {
                "success": True,
                "original_rows": original_data_rows,
                "removed_rows": removed_f1_rows,
                "final_rows": filtered_rows,
                "save_path": save_path,
                "format": export_format
            }
        
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
